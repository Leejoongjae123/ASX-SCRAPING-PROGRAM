import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication,QTreeView,QFileSystemModel,QVBoxLayout,QPushButton,QInputDialog,QLineEdit,QMainWindow,QMessageBox
from PyQt5.QtCore import QCoreApplication

from selenium.webdriver import ActionChains
from datetime import datetime,date,timedelta


import openpyxl as xl
import pyautogui as pag
import pyperclip as clp
import pyautogui
import time
import keyboard
import requests
from bs4 import BeautifulSoup as bs
import sys,os,shutil
from PyQt5.QtWidgets import QWidget, QApplication,QTreeView,QFileSystemModel,QVBoxLayout,QPushButton,QInputDialog,QLineEdit,QMainWindow
import re
import pyupbit
import pprint
import telepot
from tree_view import Ui_MainWindow





class Example(QMainWindow,Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path="C:"
        self.index=None
        self.setupUi(self)
        self.setSlot()
        self.show()
        self.progressBar.setValue(0)
        QApplication.processEvents()

    def start(self):
        # print('1')
        if self.index==None:
            QMessageBox.information(self,"경고창","파일 경로를 선택하지 않았습니다.")
            QCoreApplication.instance().quit()
        self.fpath=self.model.filePath(self.index)

        print(self.fpath)

        # # 오늘날짜 구해놓기
        #
        time_now = datetime.now()
        time_now = time_now.strftime("%d/%m/%Y")

        time_now_hour = int(datetime.now().strftime("%H"))

        # ---------------오늘 저장된 내용은 삭제하는 로직
        try:
            wb = openpyxl.load_workbook(self.fpath+'/ASX_list.xlsx')
            ws = wb.active
            no_row = ws.max_row
            print("총행의갯수:", no_row)
            for index_row in range(no_row, 1, -1):
                date_value = str(ws.cell(row=index_row, column=2).value)
                if date_value.find(time_now) >= 0:
                    print("삭제할행번호", index_row)
                    ws.delete_rows(index_row)
                else:
                    break
            wb.save(self.fpath+'/ASX_list.xlsx')
        # 파일이 없는 경우 생성
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            first_row = ['ASX CODE', 'DATE', 'PRICE SENS.', 'HEADLINE']
            ws.append(first_row)
            wb.save(self.fpath+'/ASX_list.xlsx')

        # ------------------------폴더 생성하는 함수 만들기--------------------------
        def createFolder(directory):
            try:
                if not os.path.exists(directory):
                    os.makedirs(directory)
            except OSError:
                print('Error: Creating directory. ' + directory)

        # ASX 홈페이지에 Requests 보내기
        url = 'https://www.asx.com.au/asx/v2/statistics/todayAnns.do'
        res = requests.get(url)
        soup = BeautifulSoup(res.text, 'lxml')
        table = soup.find('announcement_data')
        rows = table.find_all('tr')

        # ASX LIST 파일을 불러온다.
        wb = openpyxl.load_workbook(self.fpath+'/ASX_list.xlsx')
        ws = wb.active

        # 현재 리스트의 전체 갯수를 확인한다.
        total_list = []
        total_amount = len(rows)
        print("전체갯수:", total_amount)
        for count, row in enumerate(rows):  # 리스트 전체의 정보를 가져온다.
            if count == 0:
                continue
            each_infos = row.find_all('td')
            each_list = []
            for index, each_info in enumerate(each_infos):
                each_content = each_info.get_text()
                if index == 1:
                    each_content = each_content.strip().replace(" ", "").replace("\n", " ")
                    # print(each_content)
                if index == 2:
                    is_dollar = len(str(each_info.find('img')))
                    if is_dollar >= 10:
                        each_content = "O"
                    else:
                        each_content = "X"
                if index == 3:
                    each_content = each_content.strip()
                    each_content = each_content.split("\n")[0]

                each_list.append(each_content)
            print(each_list)
            ws.append(each_list)
            changed_time = each_list[1].replace(":", "_").replace("/", "_").replace(" ", "_")
            ASX_code = each_list[0]

            try:
                pdf_url="https://www.asx.com.au"+row.find('a')['href']
                res = requests.get(pdf_url)
                soup = BeautifulSoup(res.text, 'lxml')
                soup = str(soup)
                start = soup.find('/asx')
                soup = soup[start:]
                end = soup.find('"/>')
                soup = "https://www.asx.com.au" + soup[:end]
                print(soup)
                document_res = requests.get(soup)
                time_now = datetime.now()
                time_now = time_now.strftime("%Y%m%d")
                folder_path="/PDF_DOWNLOAD_{}_{}HOUR".format(time_now,time_now_hour)
                createFolder(self.fpath+folder_path)
                with open(self.fpath+"{}/{}_{}.pdf".format(folder_path,ASX_code,changed_time), "wb") as f:
                    f.write(document_res.content)
                    print("저장완료",total_amount,"중에서",count+1,"개 완료","({}%)".format(round(count/total_amount*100,1)))
                process=round(count/total_amount*100,1)
                self.num = int(process)
                self.progressBar.setValue(self.num)
                QApplication.processEvents()
            except:
                print("다운실패")
        wb.save(self.fpath+'/ASX_list.xlsx')
        QMessageBox.information(self, "완료창", "작업이 완료 되었습니다.")
        QCoreApplication.instance().quit()




        # for i in range(0,101):
        #     self.num=i
        #     self.progressBar.setValue(self.num)
        #     QApplication.processEvents()
        #     # QApplication.processEvents()


    def setSlot(self):
        self.treeWidget.clicked.connect(self.setIndex)

    def setIndex(self,index):
        self.index=index
    def quit(self):
        QCoreApplication.instance().quit()

app=QApplication([])
ex=Example()
sys.exit(app.exec_())



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())


















